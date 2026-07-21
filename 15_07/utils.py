from io import StringIO, BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:


    @staticmethod
    def export_employees(employees_list) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách nhân viên"

        headers = ["ID", "Họ và tên", "Email", "Số điện thoại", "Chức vụ",
                   "Phòng ban", "Lương", "Ngày vào làm", "Trạng thái"]

        # --- Style cho dòng tiêu đề ---
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # --- Ghi dữ liệu từng nhân viên ---
        for row_idx, emp in enumerate(employees_list, start=2):
            dept_name = ""
            department = getattr(emp, "department", None)
            if department is not None:
                dept_name = getattr(department, "name", "") or ""

            is_active = getattr(emp, "is_active", None)
            status_text = "Đang làm việc" if is_active else "Đã nghỉ việc"

            hire_date = getattr(emp, "hire_date", None)
            hire_date_text = hire_date.strftime("%d/%m/%Y") if hire_date else ""

            row_values = [
                getattr(emp, "id", ""),
                getattr(emp, "fullname", "") or "",
                getattr(emp, "email", "") or "",
                getattr(emp, "phone", "") or "",
                getattr(emp, "position", "") or "",
                dept_name,
                getattr(emp, "salary", None) or 0,
                hire_date_text,
                status_text,
            ]

            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 7:  # cột Lương -> canh phải, định dạng số
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")

        # --- Tự động giãn độ rộng cột theo nội dung ---
        for col_idx, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

        # Cố định dòng tiêu đề khi cuộn
        ws.freeze_panes = "A2"

        # --- Xuất ra bytes trong bộ nhớ (không tạo file tạm trên đĩa) ---
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


class CSVExporter:
    """Giữ lại CSVExporter (không dùng cho endpoint chính nữa, phòng khi cần CSV thuần)."""

    @staticmethod
    def export_employees(employees_list):
        f = StringIO()
        f.write('\ufeff')  # BOM UTF-8 cho Excel (chỉ hỗ trợ một phần, không chắc chắn 100%)
        f.write("ID,Fullname,Email,Phone,Position\n")

        def _escape(value):
            value = str(value) if value is not None else ""
            if ',' in value or '"' in value or '\n' in value:
                value = '"' + value.replace('"', '""') + '"'
            return value

        for emp in employees_list:
            emp_id = getattr(emp, 'id', '')
            fullname = getattr(emp, 'fullname', '')
            email = getattr(emp, 'email', '')
            phone = getattr(emp, 'phone', '')
            position = getattr(emp, 'position', '')
            f.write(
                f"{_escape(emp_id)},{_escape(fullname)},{_escape(email)},"
                f"{_escape(phone)},{_escape(position)}\n"
            )
        f.seek(0)
        return f


class DateFormatter:
    @staticmethod
    def get_current_datetime():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


class ResponseFormatter:
    pass